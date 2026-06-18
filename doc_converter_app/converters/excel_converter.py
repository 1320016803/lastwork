from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .converter_base import BaseConverter


class ExcelConverter(BaseConverter):
    SUPPORTED_EXT = (".xlsx", ".xls")

    def _to_markdown(self, src: Path) -> str:
        wb = load_workbook(filename=str(src), data_only=True, read_only=True)
        sections: list[str] = []

        for sheet in wb.worksheets:
            title = f"## 工作表：{sheet.title}\n"
            lines: list[list[str]] = []
            max_col = 0

            for row in sheet.iter_rows(values_only=True):
                # 跳过完全为空的行
                if row is None or all(v is None or (isinstance(v, str) and v.strip() == "") for v in row):
                    continue
                cells: list[str] = []
                for v in row:
                    if v is None:
                        cells.append("")
                    else:
                        cells.append(str(v).strip())
                lines.append(cells)
                if len(cells) > max_col:
                    max_col = len(cells)

            if not lines:
                sections.append(title + "\n（空表）\n")
                continue

            normalized = [row + [""] * (max_col - len(row)) for row in lines]
            first_row = [self._escape(c) for c in normalized[0]]
            md_lines = [
                "| " + " | ".join(first_row) + " |",
                "| " + " | ".join(["---"] * max_col) + " |",
            ]
            for row in normalized[1:]:
                md_lines.append("| " + " | ".join(self._escape(c) for c in row) + " |")

            sections.append(title + "\n" + "\n".join(md_lines) + "\n")

        wb.close()
        return "\n".join(sections).strip() + "\n"

    def _escape(self, text: str) -> str:
        return text.replace("\n", " ").replace("|", "\\|")
